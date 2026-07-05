"""
JPX高配当株 ニューロファイナンス・スコアリング分析
======================================================
・JPX東証上場全銘柄リストを取得（東証プライム・スタンダード・グロース）
・yfinanceで財務・配当データを一括取得
・7軸ニューロファイナンス・スコアで順位付け
・バックテスト（1か月前に100株投資）
・1か月先予測（線形回帰 + トレンド分解）
・結果をExcelレポートとして出力

依存パッケージ: pip install yfinance pandas numpy scikit-learn openpyxl requests
"""
from __future__ import annotations
import math, time, warnings, json, logging, sys
from datetime import date, datetime, timedelta
from pathlib import Path
import requests
import pandas as pd
import numpy as np

warnings.filterwarnings("ignore")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
log = logging.getLogger("jpx_rank")

try:
    import yfinance as yf
except ImportError:
    sys.exit("yfinance が必要です: pip install yfinance pandas numpy scikit-learn openpyxl requests")

OUTPUT_PATH = Path(r"C:\Users\ynobe\Desktop\JPX_NeuroScore_Report.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# Step 1: JPX東証上場銘柄リスト取得
# ─────────────────────────────────────────────────────────────────────────────
# JPXが公開するCSV（東証上場銘柄一覧）を取得する
# URL: https://www.jpx.co.jp/markets/statistics-equities/misc/01.html から
# 直接ファイルリンクを取得

JPX_CSV_URL = (
    "https://www.jpx.co.jp/markets/statistics-equities/misc/tvdivq0000001vg2-att/"
    "data_j.xls"
)

def fetch_jpx_ticker_list() -> pd.DataFrame:
    """東証上場銘柄一覧をJPXサイトから取得する。
    取得失敗時はハードコードのサンプルリスト（主要高配当銘柄）で代替。
    """
    log.info("JPX銘柄リストを取得中...")
    try:
        r = requests.get(JPX_CSV_URL, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code == 200:
            from io import BytesIO
            df = pd.read_excel(BytesIO(r.content), header=0)
            df.columns = [str(c).strip() for c in df.columns]
            # コード列を検索
            code_col = next((c for c in df.columns if "コード" in c or "code" in c.lower()), None)
            name_col = next((c for c in df.columns if "銘柄名" in c or "会社名" in c), None)
            market_col = next((c for c in df.columns if "市場" in c or "区分" in c), None)
            if code_col:
                result = pd.DataFrame()
                result["code"] = df[code_col].astype(str).str.strip().str.zfill(4)
                result["name"] = df[name_col].astype(str) if name_col else result["code"]
                result["market"] = df[market_col].astype(str) if market_col else "東証"
                result = result[result["code"].str.match(r"^\d{4}$")]
                log.info(f"  → {len(result)} 銘柄取得（JPX公式）")
                return result
    except Exception as e:
        log.warning(f"JPX CSV取得失敗: {e} → ローカルサンプルで代替")

    # フォールバック: 東証プライム主要高配当銘柄 200銘柄
    SAMPLE_TICKERS = [
        # 銀行
        ("8306","三菱UFJ"), ("8316","三井住友"), ("8411","みずほ"), ("8354","ふくおか"),
        ("8308","りそな"), ("8309","三井住友ト"), ("8331","千葉銀行"), ("8332","横浜銀行"),
        ("8337","千葉興業"), ("8341","七十七銀"), ("8356","十六銀行"), ("8360","山梨中央"),
        ("8366","滋賀銀行"), ("8370","京都銀行"), ("8374","四国銀行"), ("8379","広島銀行"),
        ("8385","伊予銀行"), ("8386","百十四銀"), ("8387","四国銀行"), ("8393","宮崎銀行"),
        # 通信
        ("9432","NTT"), ("9433","KDDI"), ("9434","ソフトバンク"), ("9437","NTTドコモ"),
        # インフラ・公益
        ("9501","東京電力"), ("9502","中部電力"), ("9503","関西電力"), ("9531","東京ガス"),
        ("9532","大阪ガス"), ("9533","東邦ガス"), ("9505","北陸電力"),
        # 商社
        ("8058","三菱商事"), ("8031","三井物産"), ("8001","伊藤忠"), ("8002","丸紅"),
        ("8053","住友商事"), ("8056","日立ハイテク"), ("9010","富士急"),
        # タバコ・飲料
        ("2914","日本たばこ"), ("2503","キリン"), ("2502","アサヒ"), ("2587","サントリー"),
        # 石油・化学
        ("5019","出光興産"), ("5020","ENEOSホールディングス"), ("4005","住友化学"),
        ("4183","三井化学"), ("4188","三菱ケミカル"), ("4452","花王"),
        # 鉄鋼・非鉄
        ("5401","日本製鉄"), ("5411","JFEホールディングス"), ("5714","DOWAホールディングス"),
        ("5802","住友電工"), ("5803","フジクラ"),
        # 自動車・機械
        ("7203","トヨタ"), ("7201","日産"), ("7202","いすゞ"), ("7269","スズキ"),
        ("7270","SUBARU"), ("7272","ヤマハ発"), ("6301","コマツ"), ("6302","住友重機"),
        # 不動産
        ("8801","三井不動産"), ("8802","三菱地所"), ("8830","住友不動産"),
        ("3003","ヒューリック"), ("3234","森ヒルズ"), ("8815","東急不動産"),
        # 保険
        ("8750","第一生命"), ("8752","三井住友海上"), ("8766","東京海上"),
        ("8795","太陽生命"), ("8804","東京建物"),
        # 小売
        ("8267","イオン"), ("8270","ユニー"), ("9843","ニトリ"), ("3382","セブン&アイ"),
        ("2651","ローソン"), ("2706","ブロンコビリー"),
        # IT・電機
        ("6758","ソニー"), ("6501","日立"), ("6504","富士電機"), ("6508","明電舎"),
        ("6701","NEC"), ("6702","富士通"), ("6703","沖電気"), ("6752","パナソニック"),
        ("6753","シャープ"), ("6764","三洋電機"),
        # その他高配当
        ("4502","武田薬品"), ("4503","アステラス"), ("4506","大日本住友"),
        ("4523","エーザイ"), ("4568","第一三共"),
        ("6952","カシオ"), ("7012","川崎重工"), ("7013","IHI"), ("7011","三菱重工"),
        ("8002","丸紅"), ("5706","三菱マテリアル"), ("5713","住友鉱山"),
        ("1803","清水建設"), ("1802","大林組"), ("1801","大成建設"),
        ("4901","富士フイルム"), ("4902","コニカミノルタ"),
        ("9601","松竹"), ("9602","東宝"), ("9603","エイチ・アイ・エス"),
        ("7951","ヤマハ"), ("5411","JFEホールディングス"), ("6361","荏原製作所"),
        ("6366","千代田化工"),("6367","ダイキン"), ("6471","日本精工"), ("6472","NTN"),
        ("6473","ジェイテクト"), ("6478","THK"), ("6501","日立"),
        ("7741","HOYA"), ("7762","シチズン"), ("7751","キヤノン"),
        ("5901","東洋製罐"), ("5902","東罐興業"), ("5903","ベイシア"),
        ("2802","味の素"), ("2801","キッコーマン"), ("2871","ニッスイ"), ("2872","セイヒョー"),
        ("3101","東洋紡"), ("3103","ユニチカ"), ("3401","帝人"), ("3402","東レ"),
        ("3861","王子HD"), ("3863","日本製紙"), ("3864","三菱製紙"),
        ("9001","東武鉄道"), ("9002","西武HD"), ("9005","東急"),
        ("9006","京浜急行"), ("9007","小田急"), ("9008","京王"),
        ("9009","京成"), ("9020","東日本旅客鉄道"), ("9021","西日本旅客鉄道"),
    ]
    df = pd.DataFrame(SAMPLE_TICKERS, columns=["code", "name"])
    df["market"] = "東証プライム"
    # 重複除去
    df = df.drop_duplicates("code")
    log.info(f"  → サンプルリスト {len(df)} 銘柄を使用")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Step 2: yfinanceで株価・配当データ取得
# ─────────────────────────────────────────────────────────────────────────────

def fetch_stock_data(tickers_df: pd.DataFrame, max_tickers: int = 300) -> pd.DataFrame:
    """yfinanceで各銘柄の財務データを取得する。
    日本株のtickerは '1234.T' 形式。
    """
    codes = tickers_df["code"].tolist()[:max_tickers]
    yf_tickers = [f"{c}.T" for c in codes]

    log.info(f"{len(yf_tickers)} 銘柄のデータを取得中（バッチ処理）...")

    end_date = datetime.today()
    start_date = end_date - timedelta(days=400)  # 1年強

    records = []
    batch_size = 20

    for i in range(0, len(yf_tickers), batch_size):
        batch = yf_tickers[i:i+batch_size]
        batch_codes = codes[i:i+batch_size]
        log.info(f"  バッチ {i//batch_size + 1}/{math.ceil(len(yf_tickers)/batch_size)}: {batch[0]}〜{batch[-1]}")

        for j, (yticker, code) in enumerate(zip(batch, batch_codes)):
            name = tickers_df[tickers_df["code"]==code]["name"].iloc[0] if code in tickers_df["code"].values else code
            try:
                t = yf.Ticker(yticker)
                info = t.fast_info  # 軽量版

                # 株価
                price = getattr(info, "last_price", None) or getattr(info, "regularMarketPrice", None)
                if not price or price <= 0:
                    continue

                # 詳細情報（遅いが一度だけ）
                try:
                    full_info = t.info
                except:
                    full_info = {}

                # 配当利回り
                div_yield = full_info.get("dividendYield", 0) or 0
                # 1株配当 (DPS)
                div_rate = full_info.get("dividendRate", 0) or 0  # 年間配当（USD/JPY）
                # 配当性向
                payout_ratio = full_info.get("payoutRatio", 0) or 0
                # 自己資本比率
                # yfinanceは直接提供しないため、BS/PLから計算
                equity_ratio = 0.30  # デフォルト
                debt_equity = 1.0    # デフォルト
                try:
                    bs = t.balance_sheet
                    if bs is not None and not bs.empty:
                        te_keys = [k for k in bs.index if "Stockholders" in str(k) or "Equity" in str(k)]
                        ta_keys = [k for k in bs.index if "Total Assets" in str(k) or "TotalAssets" in str(k)]
                        tl_keys = [k for k in bs.index if "Total Liab" in str(k) or "TotalLiab" in str(k)]
                        ltd_keys = [k for k in bs.index if "Long Term Debt" in str(k) or "LongTermDebt" in str(k)]

                        if te_keys and ta_keys:
                            te = float(bs.loc[te_keys[0]].iloc[0])
                            ta = float(bs.loc[ta_keys[0]].iloc[0])
                            if ta > 0:
                                equity_ratio = te / ta

                        if te_keys and ltd_keys:
                            te_v = float(bs.loc[te_keys[0]].iloc[0])
                            ltd_v = float(bs.loc[ltd_keys[0]].iloc[0])
                            if te_v > 0:
                                debt_equity = ltd_v / te_v
                except:
                    pass

                # 配当履歴
                dps_history = []
                consecutive_raises = 0
                has_cut = False
                try:
                    divs = t.dividends
                    if divs is not None and len(divs) > 0:
                        # 年次集計
                        div_annual = divs.groupby(divs.index.year).sum()
                        dps_history = div_annual.tolist()
                        # 連続増配
                        n_raises = 0
                        for k in range(len(dps_history)-1, 0, -1):
                            if dps_history[k] > dps_history[k-1] * 1.001:
                                n_raises += 1
                            else:
                                break
                        consecutive_raises = n_raises
                        # 減配チェック
                        for k in range(1, len(dps_history)):
                            if dps_history[k] < dps_history[k-1] * 0.95:
                                has_cut = True
                                break
                except:
                    pass

                # DPS 補完
                if div_rate <= 0 and div_yield > 0:
                    div_rate = price * div_yield
                if div_rate <= 0 and dps_history:
                    div_rate = dps_history[-1] if dps_history else 0

                # payout_ratio補完
                if payout_ratio <= 0 or payout_ratio > 1.5:
                    eps = full_info.get("trailingEps", 0) or 0
                    if eps > 0 and div_rate > 0:
                        payout_ratio = div_rate / eps
                    else:
                        payout_ratio = 0.50

                market_cap_m = (full_info.get("marketCap", 0) or 0) / 1_000_000
                sector = full_info.get("sector", "") or full_info.get("industry", "") or ""

                records.append({
                    "ticker": code,
                    "name": name,
                    "price": price,
                    "dps": round(div_rate, 1),
                    "div_yield_pct": round(div_yield * 100, 2),
                    "payout_ratio": min(payout_ratio, 1.5),
                    "equity_ratio": min(max(equity_ratio, 0), 1.0),
                    "debt_equity": max(debt_equity, 0),
                    "consecutive_raises": consecutive_raises,
                    "has_dividend_cut": has_cut,
                    "dps_history": dps_history,
                    "market_cap_m": market_cap_m,
                    "sector": sector,
                })

            except Exception as e:
                log.debug(f"  スキップ {yticker}: {e}")
                continue

            time.sleep(0.25)  # レート制限

        time.sleep(1.0)

    df = pd.DataFrame(records)
    log.info(f"取得完了: {len(df)} 銘柄（有効データあり）")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Step 3: 7軸ニューロファイナンス・スコアリング
# ─────────────────────────────────────────────────────────────────────────────

LAMBDA = 2.0  # Tom et al. (2007) 損失回避係数

WEIGHTS = {
    "stability":   0.25,  # [P3] Insula 独立回路
    "health":      0.20,  # [P1] Insula 誘発予防
    "yield":       0.20,  # [P3] NAcc 期待報酬（上限付き）
    "momentum":    0.15,  # [P4] ドーパミン予測誤差
    "payout":      0.10,  # [P2] λ=2 過剰性向ペナルティ
    "streak":      0.07,  # [P4] 連続正予測誤差
    "sector_rank": 0.03,  # 社会比較参照点
}

def _cv(history: list) -> float:
    """変動係数 (Insula活性の代理指標)"""
    vals = [v for v in history if v > 0]
    if len(vals) < 2:
        return 0.5
    mean = np.mean(vals)
    return np.std(vals) / mean if mean > 0 else 1.0

def _cagr_3y(history: list) -> float:
    """直近3年配当CAGR (ドーパミン予測誤差)"""
    vals = [v for v in history if v > 0]
    if len(vals) < 2:
        return 0.0
    n = min(3, len(vals) - 1)
    if vals[-(n+1)] <= 0:
        return 0.0
    return (vals[-1] / vals[-(n+1)]) ** (1.0/n) - 1.0

def score_stability(history: list, has_cut: bool) -> float:
    """[P3][P5] Insula 活性化抑制"""
    cv = _cv(history)
    if cv <= 0.10:
        base = 1.0
    elif cv <= 0.50:
        base = 1.0 - (cv - 0.10) / 0.40
    else:
        base = 0.0
    if has_cut:
        return min(base * 0.10, 0.05)
    return max(0.0, base)

def score_health(equity_ratio: float, debt_equity: float) -> float:
    """[P1][P2] Insula誘発予防 + λ=2"""
    if equity_ratio >= 0.40:
        eq_s = 1.0
    elif equity_ratio >= 0.20:
        eq_s = (equity_ratio - 0.20) / 0.20 * 0.7
    else:
        eq_s = equity_ratio / 0.20 * 0.20

    if debt_equity <= 1.0:
        de_s = 1.0 - debt_equity * 0.33
    elif debt_equity <= 3.0:
        de_s = max(0, 0.67 - (debt_equity - 1.0) / 2.0 * 0.67)
    else:
        de_s = 0.0
    return max(0.0, eq_s * 0.55 + de_s * 0.45)

def score_yield(dy: float) -> float:
    """[P1][P3] NAcc 過活性化防止 — 高利回りに上限"""
    if dy <= 0:
        return 0.0
    if dy > 0.12:
        return 0.10
    if dy > 0.10:
        return max(0.10, 0.30 - (dy - 0.10) / 0.02 * 0.20)
    if dy > 0.08:
        return max(0.10, 0.80 - (dy - 0.08) / 0.02 * 0.50)
    return min(1.0, 1 / (1 + math.exp(-15 * (dy - 0.04))))

def score_momentum(history: list) -> float:
    """[P4] ドーパミン予測誤差 — 直近3年CAGR"""
    if len(history) < 2:
        return 0.35
    cagr = _cagr_3y(history)
    if cagr >= 0.10:
        return 1.0
    if cagr >= 0:
        return 0.40 + cagr / 0.10 * 0.60
    return max(0.0, 0.40 + cagr / 0.10 * 0.60 * LAMBDA)

def score_payout(ratio: float) -> float:
    """[P2] λ=2 非対称ペナルティ"""
    if ratio <= 0:
        return 0.20
    if ratio > 1.20:
        return 0.0
    if ratio > 1.0:
        return max(0.0, 0.10 - (ratio - 1.0) / 0.20 * 0.10)
    if ratio > 0.80:
        return max(0.0, 1.0 - (ratio - 0.80) / 0.20 * 0.90)
    if 0.30 <= ratio <= 0.70:
        return 1.0
    if ratio < 0.30:
        return 0.20 + (ratio / 0.30) * 0.80
    return max(0.0, 1.0 - (ratio - 0.70) / 0.10 * 0.10)

def score_streak(years: int) -> float:
    """[P4] 連続正予測誤差の累積"""
    if years <= 0:
        return 0.0
    return min(1.0, math.log(years + 1) / math.log(26))

def compute_scores(df: pd.DataFrame) -> pd.DataFrame:
    """全銘柄スコアリング"""
    log.info("スコア計算中...")

    results = []
    for _, row in df.iterrows():
        history = row["dps_history"] if isinstance(row["dps_history"], list) else []
        dy = row["div_yield_pct"] / 100.0 if row["div_yield_pct"] > 0 else (
            row["dps"] / row["price"] if row["price"] > 0 and row["dps"] > 0 else 0
        )

        s_stab = score_stability(history, row["has_dividend_cut"])
        s_hlth = score_health(row["equity_ratio"], row["debt_equity"])
        s_yld  = score_yield(dy)
        s_mom  = score_momentum(history)
        s_pay  = score_payout(row["payout_ratio"])
        s_str  = score_streak(row["consecutive_raises"])
        s_sect = 0.50  # sector rank — 後でパーセンタイル計算

        cv_val = _cv(history)
        cagr_val = _cagr_3y(history)

        results.append({
            **row.to_dict(),
            "div_yield_pct": round(dy * 100, 2),
            "s_stability": round(s_stab, 4),
            "s_health":    round(s_hlth, 4),
            "s_yield":     round(s_yld, 4),
            "s_momentum":  round(s_mom, 4),
            "s_payout":    round(s_pay, 4),
            "s_streak":    round(s_str, 4),
            "s_sector_rank": round(s_sect, 4),
            "dps_cv": round(cv_val, 3),
            "dps_cagr_3y_pct": round(cagr_val * 100, 1),
        })

    scored = pd.DataFrame(results)

    # Sector rank: 利回りの同業種内パーセンタイルで更新
    for _, grp in scored.groupby("sector"):
        if len(grp) > 1:
            rank_vals = grp["div_yield_pct"].rank(pct=True, ascending=False)
            scored.loc[grp.index, "s_sector_rank"] = rank_vals.apply(
                lambda r: round(1.0 - r, 4)
            )

    # 総合スコア計算
    w = WEIGHTS
    scored["total_score"] = (
        scored["s_stability"]   * w["stability"] +
        scored["s_health"]      * w["health"] +
        scored["s_yield"]       * w["yield"] +
        scored["s_momentum"]    * w["momentum"] +
        scored["s_payout"]      * w["payout"] +
        scored["s_streak"]      * w["streak"] +
        scored["s_sector_rank"] * w["sector_rank"]
    ).round(4)

    # 配当ありのみフィルタ（div_yield > 0）、スコア順にソート
    scored = scored[scored["div_yield_pct"] > 0].copy()
    scored = scored.sort_values("total_score", ascending=False).reset_index(drop=True)
    scored["rank"] = range(1, len(scored) + 1)

    log.info(f"スコアリング完了: {len(scored)} 銘柄（配当あり）")
    return scored


# ─────────────────────────────────────────────────────────────────────────────
# Step 4: バックテスト（1か月前に100株投資）
# ─────────────────────────────────────────────────────────────────────────────

def run_backtest(top_df: pd.DataFrame, top_n: int = 20) -> pd.DataFrame:
    """上位N銘柄に1か月前（約21営業日前）に100株投資した場合の損益。"""
    log.info(f"バックテスト実行中（上位 {top_n} 銘柄）...")

    end_date = datetime.today()
    start_date = end_date - timedelta(days=60)  # 2か月分（余裕を持って）

    bt_records = []
    for _, row in top_df.head(top_n).iterrows():
        ticker_code = row["ticker"]
        yticker = f"{ticker_code}.T"
        try:
            hist = yf.download(yticker, start=start_date, end=end_date,
                              progress=False, auto_adjust=True)
            if hist is None or hist.empty:
                continue

            close = hist["Close"].dropna()
            if len(close) < 10:
                continue

            # 1か月前（約21営業日前）の価格
            idx_1m_ago = max(0, len(close) - 22)
            price_1m_ago = float(close.iloc[idx_1m_ago])
            price_now = float(close.iloc[-1])

            shares = 100
            cost = price_1m_ago * shares
            value_now = price_now * shares
            pnl = value_now - cost
            pnl_pct = pnl / cost * 100

            # 配当（1か月分: 年間配当 / 12）
            annual_div = row["dps"]
            monthly_div = annual_div / 12 * shares
            total_return = pnl + monthly_div
            total_return_pct = total_return / cost * 100

            bt_records.append({
                "rank": row["rank"],
                "ticker": ticker_code,
                "name": row["name"],
                "total_score": row["total_score"],
                "div_yield_pct": row["div_yield_pct"],
                "price_1m_ago": round(price_1m_ago, 1),
                "price_now": round(price_now, 1),
                "shares": shares,
                "cost_jpy": round(cost, 0),
                "value_now_jpy": round(value_now, 0),
                "pnl_jpy": round(pnl, 0),
                "monthly_div_jpy": round(monthly_div, 1),
                "total_return_jpy": round(total_return, 0),
                "price_change_pct": round(pnl_pct, 2),
                "total_return_pct": round(total_return_pct, 2),
            })
            time.sleep(0.5)
        except Exception as e:
            log.debug(f"バックテストスキップ {ticker_code}: {e}")

    bt_df = pd.DataFrame(bt_records)
    if not bt_df.empty:
        log.info(f"バックテスト完了: {len(bt_df)} 銘柄")
        avg_return = bt_df["total_return_pct"].mean()
        winners = (bt_df["total_return_pct"] > 0).sum()
        log.info(f"  平均リターン: {avg_return:.2f}%、勝率: {winners}/{len(bt_df)}")
    return bt_df


# ─────────────────────────────────────────────────────────────────────────────
# Step 5: 1か月先価格予測（線形回帰 + トレンド）
# ─────────────────────────────────────────────────────────────────────────────

def predict_1month(top_df: pd.DataFrame, top_n: int = 20) -> pd.DataFrame:
    """上位N銘柄の1か月先株価を線形回帰で予測する。"""
    from sklearn.linear_model import LinearRegression

    log.info(f"1か月先予測中（上位 {top_n} 銘柄）...")

    end_date = datetime.today()
    start_date = end_date - timedelta(days=365)

    pred_records = []
    for _, row in top_df.head(top_n).iterrows():
        ticker_code = row["ticker"]
        yticker = f"{ticker_code}.T"
        try:
            hist = yf.download(yticker, start=start_date, end=end_date,
                              progress=False, auto_adjust=True)
            if hist is None or hist.empty or len(hist) < 30:
                continue

            close = hist["Close"].dropna().values
            X = np.arange(len(close)).reshape(-1, 1)
            y = close

            # 線形トレンド回帰
            lr = LinearRegression()
            lr.fit(X, y)

            # 1か月先（約21営業日）
            X_pred = np.array([[len(close) + 21]])
            price_pred = float(lr.predict(X_pred)[0])

            # 信頼区間の近似（残差の標準偏差）
            y_hat = lr.predict(X)
            residuals = y - y_hat
            sigma = np.std(residuals)

            price_now = float(close[-1])
            change_pct = (price_pred - price_now) / price_now * 100

            # スコアによるモメンタム補正（高スコア銘柄に微小な上乗せ）
            # ただしこれはモデルの仮定であり、保証ではない
            neuro_adj = (row["total_score"] - 0.5) * 0.02  # ±1%の補正
            price_pred_adj = price_pred * (1 + neuro_adj)
            change_pct_adj = (price_pred_adj - price_now) / price_now * 100

            # リスク指標: 価格の30日ボラティリティ（年率）
            if len(close) > 30:
                returns = np.diff(np.log(close[-31:]))
                vol_annual = np.std(returns) * np.sqrt(252) * 100
            else:
                vol_annual = np.nan

            pred_records.append({
                "rank": row["rank"],
                "ticker": ticker_code,
                "name": row["name"],
                "total_score": row["total_score"],
                "div_yield_pct": row["div_yield_pct"],
                "price_now": round(price_now, 1),
                "price_pred_1m": round(price_pred_adj, 1),
                "change_pct_1m": round(change_pct_adj, 2),
                "upper_1sigma": round(price_pred_adj + sigma, 1),
                "lower_1sigma": round(price_pred_adj - sigma, 1),
                "trend_slope_daily": round(float(lr.coef_[0]), 2),
                "vol_annual_pct": round(vol_annual, 1) if not np.isnan(vol_annual) else None,
                "model": "線形回帰 + ニューロスコア補正",
            })
            time.sleep(0.5)
        except Exception as e:
            log.debug(f"予測スキップ {ticker_code}: {e}")

    pred_df = pd.DataFrame(pred_records)
    log.info(f"予測完了: {len(pred_df)} 銘柄")
    return pred_df


# ─────────────────────────────────────────────────────────────────────────────
# Step 6: Excelレポート出力
# ─────────────────────────────────────────────────────────────────────────────

def write_excel_report(scored: pd.DataFrame, backtest: pd.DataFrame,
                       prediction: pd.DataFrame, output_path: Path):
    """投資家向けExcelレポートを生成する。"""
    log.info(f"Excelレポート生成中: {output_path}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                     numbers)
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.chart.series import SeriesLabel
    except ImportError:
        log.error("openpyxl が必要: pip install openpyxl")
        return

    wb = Workbook()

    # カラーパレット
    C_HEADER   = "1E3A5F"  # ダークネイビー
    C_SUBHEAD  = "2563EB"  # ブルー
    C_GOLD     = "F59E0B"  # ゴールド（トップ銘柄）
    C_GREEN    = "16A34A"  # 緑（プラス）
    C_RED      = "DC2626"  # 赤（マイナス）
    C_LIGHT    = "EFF6FF"  # ライトブルー
    C_WHITE    = "FFFFFF"
    C_GRAY     = "F3F4F6"
    C_BORDER   = "CBD5E1"

    thin = Side(border_style="thin", color=C_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def h_fill(color): return PatternFill("solid", fgColor=color)
    def h_font(color=C_WHITE, bold=True, size=10):
        return Font(name="Meiryo UI", color=color, bold=bold, size=size)
    def align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ────────────── Sheet 1: 表紙 ──────────────
    ws0 = wb.active
    ws0.title = "📊表紙"
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 3
    ws0.column_dimensions["B"].width = 60
    ws0.column_dimensions["C"].width = 20

    ws0.row_dimensions[1].height = 20
    ws0.row_dimensions[2].height = 50
    ws0.row_dimensions[3].height = 40

    ws0.merge_cells("B2:C2")
    c = ws0["B2"]
    c.value = "JPX高配当株 ニューロファイナンス・スコアリング レポート"
    c.font = Font(name="Meiryo UI", size=20, bold=True, color=C_HEADER)
    c.alignment = align("left")
    c.fill = h_fill(C_WHITE)

    ws0.merge_cells("B3:C3")
    ws0["B3"].value = f"作成日: {date.today().strftime('%Y年%m月%d日')}　　分析対象: {len(scored)} 銘柄（配当あり）"
    ws0["B3"].font = Font(name="Meiryo UI", size=11, color="6B7280")
    ws0["B3"].alignment = align("left")

    info_rows = [
        ("", ""),
        ("■ 分析フレームワーク", ""),
        ("スコアリングモデル", "7軸ニューロファイナンス（神経科学論文に基づく）"),
        ("", ""),
        ("📐 重み配分 (7軸)", ""),
        ("  安定性 (Stability)", f"25% — Knutson & Bossaerts (2007) Insula独立回路"),
        ("  財務健全性 (Health)", f"20% — Kuhnen & Knutson (2005) Insula誘発予防"),
        ("  利回り (Yield)", f"20% — NAcc期待報酬符号化（過剰比重を回避）"),
        ("  成長性 (Momentum)", f"15% — Schultz et al. (1997) ドーパミン予測誤差"),
        ("  配当性向 (Payout)", f"10% — Tom et al. (2007) λ=2 損失回避ペナルティ"),
        ("  連続増配 (Streak)", f"7%  — ドーパミン連続正予測誤差の累積"),
        ("  業種内順位 (Sector)", f"3%  — 社会比較効果（Fliessbach et al. 2007）"),
        ("", ""),
        ("⚠ 注意事項", ""),
        ("本レポートは", "投資判断の参考情報です。投資損益の責任はご自身で負ってください。"),
        ("データソース", "Yahoo Finance Japan（yfinance経由）、JPX上場銘柄一覧"),
        ("予測モデル", "線形回帰 + ニューロスコア補正（過去トレンドを前提とした統計的推定）"),
    ]
    for r, (k, v) in enumerate(info_rows, start=5):
        ws0.row_dimensions[r].height = 18
        ck = ws0.cell(r, 2, value=k)
        cv = ws0.cell(r, 3, value=v)
        if k.startswith("■") or k.startswith("📐") or k.startswith("⚠"):
            ck.font = Font(name="Meiryo UI", bold=True, size=10, color=C_HEADER)
        elif k.startswith("  "):
            ck.font = Font(name="Meiryo UI", size=10, color="374151")
            cv.font = Font(name="Meiryo UI", size=10, color="374151")
        else:
            ck.font = Font(name="Meiryo UI", size=10, color="1F2937")
            cv.font = Font(name="Meiryo UI", size=10, color="374151")
        ck.alignment = align("left")
        cv.alignment = align("left", wrap=True)

    # ────────────── Sheet 2: 全銘柄ランキング ──────────────
    ws1 = wb.create_sheet("🏆スコアランキング")
    ws1.sheet_view.showGridLines = False

    cols = [
        ("順位", 6), ("コード", 8), ("銘柄名", 22), ("セクター", 16),
        ("利回り%", 9), ("連配年", 8), ("性向%", 8), ("自己資本%", 9),
        ("安定性", 8), ("財務", 8), ("利回得点", 8), ("成長性", 8),
        ("性向得点", 8), ("連配得点", 8), ("業種内", 8), ("総合スコア", 10),
        ("減配歴", 8),
    ]
    for cidx, (label, width) in enumerate(cols, start=1):
        ws1.column_dimensions[get_column_letter(cidx)].width = width
        c = ws1.cell(1, cidx, value=label)
        c.fill = h_fill(C_HEADER)
        c.font = h_font()
        c.alignment = align()
        c.border = border

    ws1.row_dimensions[1].height = 24

    score_col = len(cols) - 1  # 総合スコア列（最終列「減配歴」の1つ前）
    for ridx, row in scored.iterrows():
        r = ridx + 2
        rank = row["rank"]
        row_vals = [
            rank, row["ticker"], row["name"], row.get("sector", ""),
            row["div_yield_pct"], row["consecutive_raises"],
            round(row["payout_ratio"]*100, 1), round(row["equity_ratio"]*100, 1),
            row["s_stability"], row["s_health"], row["s_yield"], row["s_momentum"],
            row["s_payout"], row["s_streak"], row["s_sector_rank"], row["total_score"],
            "あり⚠" if row["has_dividend_cut"] else "なし",
        ]
        is_top3 = rank <= 3
        is_top10 = rank <= 10
        for cidx, val in enumerate(row_vals, start=1):
            c = ws1.cell(r, cidx, value=val)
            c.border = border
            c.alignment = align("center" if cidx != 3 else "left")
            c.font = Font(name="Meiryo UI", size=9,
                         bold=is_top3,
                         color=C_HEADER if is_top3 else ("374151" if is_top10 else "6B7280"))
            if is_top3:
                c.fill = h_fill("FEF3C7")  # ゴールド系
            elif is_top10:
                c.fill = h_fill(C_LIGHT)
            elif ridx % 2 == 0:
                c.fill = h_fill(C_GRAY)
            # 総合スコアのカラーバー的な強調
            if cidx == score_col:
                try:
                    score_val = float(val) if val else 0
                except (ValueError, TypeError):
                    score_val = 0
                if score_val >= 0.70:
                    c.fill = h_fill("DCFCE7")
                    c.font = Font(name="Meiryo UI", size=9, bold=True, color="15803D")
                elif score_val >= 0.60:
                    c.fill = h_fill("DBEAFE")
                    c.font = Font(name="Meiryo UI", size=9, bold=True, color="1D4ED8")
            # 減配ありは赤
            if cidx == len(cols) and val == "あり⚠":
                c.font = Font(name="Meiryo UI", size=9, color=C_RED, bold=True)

        ws1.row_dimensions[r].height = 16

    # ────────────── Sheet 3: 上位20銘柄詳細 ──────────────
    ws2 = wb.create_sheet("🔍上位20銘柄詳細")
    ws2.sheet_view.showGridLines = False

    top20 = scored.head(20)
    detail_cols = [
        ("順位", 6), ("コード", 8), ("銘柄名", 24),
        ("株価(円)", 10), ("1株配当", 9), ("利回%", 8),
        ("3年CAGR%", 9), ("配当CV", 8), ("連配(年)", 8), ("性向%", 8),
        ("自己資本%", 9), ("D/Eﾚｼｵ", 8),
        ("安定性", 8), ("財務", 8), ("利回得点", 8), ("成長性", 8),
        ("性向得点", 8), ("連配得点", 8), ("業種内", 8), ("総合スコア", 10),
        ("主なリスク", 32),
    ]
    for cidx, (label, width) in enumerate(detail_cols, start=1):
        ws2.column_dimensions[get_column_letter(cidx)].width = width
        c = ws2.cell(1, cidx, value=label)
        c.fill = h_fill(C_SUBHEAD)
        c.font = h_font()
        c.alignment = align()
        c.border = border
    ws2.row_dimensions[1].height = 26

    for ridx, row in top20.iterrows():
        r = ridx + 2  # top20はindex 0始まりなので+2

        # リスク文章生成
        risks = []
        if row["has_dividend_cut"]:
            risks.append("[P5] 減配履歴あり: 投資家心理の毀損リスク")
        if row["div_yield_pct"] > 8:
            risks.append(f"[P1] 利回り{row['div_yield_pct']:.1f}%: NAcc過活性ゾーン（株価下落リスク）")
        if row["payout_ratio"] > 0.80:
            risks.append(f"[P2] 性向{row['payout_ratio']*100:.0f}%: 将来削減リスク高")
        if row["dps_cv"] > 0.30:
            risks.append(f"[P3] 変動係数{row['dps_cv']:.2f}: Insula活性帯（保有継続が難しい）")
        if not risks:
            risks.append("主要リスク指標は許容範囲内")

        row_vals = [
            row["rank"], row["ticker"], row["name"],
            row["price"], row["dps"], row["div_yield_pct"],
            row["dps_cagr_3y_pct"], row["dps_cv"], row["consecutive_raises"],
            round(row["payout_ratio"]*100, 1), round(row["equity_ratio"]*100, 1),
            row["debt_equity"],
            row["s_stability"], row["s_health"], row["s_yield"], row["s_momentum"],
            row["s_payout"], row["s_streak"], row["s_sector_rank"], row["total_score"],
            " / ".join(risks),
        ]
        for cidx, val in enumerate(row_vals, start=1):
            c = ws2.cell(r, cidx, value=val)
            c.border = border
            c.alignment = align("left" if cidx in [3, len(detail_cols)] else "center",
                               wrap=True if cidx == len(detail_cols) else False)
            c.font = Font(name="Meiryo UI", size=9)
            if ridx < 3:
                c.fill = h_fill("FEF3C7")
                c.font = Font(name="Meiryo UI", size=9, bold=True)
            elif ridx % 2 == 0:
                c.fill = h_fill(C_GRAY)
        ws2.row_dimensions[r].height = 28

    # ────────────── Sheet 4: バックテスト結果 ──────────────
    ws3 = wb.create_sheet("📈バックテスト(1か月)")
    ws3.sheet_view.showGridLines = False

    # サマリー
    if not backtest.empty:
        avg_ret = backtest["total_return_pct"].mean()
        win_rate = (backtest["total_return_pct"] > 0).mean() * 100
        total_cost = backtest["cost_jpy"].sum()
        total_val = backtest["value_now_jpy"].sum()
        total_div = backtest["monthly_div_jpy"].sum()
        portfolio_ret = (total_val + total_div - total_cost) / total_cost * 100

        summary_rows = [
            ("📊 バックテストサマリー（上位20銘柄 × 100株 = ポートフォリオ）", ""),
            ("", ""),
            ("分析期間", f"約1か月（{(datetime.today()-timedelta(days=22)).strftime('%Y/%m/%d')} → {datetime.today().strftime('%Y/%m/%d')}）"),
            ("投資戦略", f"ニューロスコア上位20銘柄に各100株を均等投資"),
            ("総投資額", f"¥{total_cost:,.0f}"),
            ("現在評価額", f"¥{total_val:,.0f}"),
            ("受取配当(月間)", f"¥{total_div:,.0f}"),
            ("ポートフォリオ総リターン", f"{portfolio_ret:+.2f}%"),
            ("平均リターン（1銘柄）", f"{avg_ret:+.2f}%"),
            ("勝率（プラスリターン）", f"{win_rate:.1f}%"),
        ]
        for ri, (k, v) in enumerate(summary_rows, start=1):
            ws3.row_dimensions[ri].height = 20
            ck = ws3.cell(ri, 1, value=k)
            cv = ws3.cell(ri, 2, value=v)
            if "サマリー" in k:
                ck.font = Font(name="Meiryo UI", bold=True, size=12, color=C_HEADER)
                ws3.merge_cells(f"A{ri}:E{ri}")
            else:
                ck.font = Font(name="Meiryo UI", bold=True, size=10, color="374151")
                cv.font = Font(name="Meiryo UI", size=10, color="374151")
            ck.alignment = align("left")
            cv.alignment = align("left")

        bt_header_row = len(summary_rows) + 2
        bt_cols = [
            ("スコア順位", 8), ("コード", 8), ("銘柄名", 22), ("総合スコア", 10),
            ("利回%", 8), ("1か月前株価", 11), ("現在株価", 10), ("株価変化%", 9),
            ("投資額(¥)", 11), ("現在評価(¥)", 11), ("株価損益(¥)", 11),
            ("月間配当(¥)", 10), ("総リターン(¥)", 11), ("総リターン%", 10),
        ]
        for cidx, (label, width) in enumerate(bt_cols, start=1):
            ws3.column_dimensions[get_column_letter(cidx)].width = width
            c = ws3.cell(bt_header_row, cidx, value=label)
            c.fill = h_fill(C_HEADER)
            c.font = h_font()
            c.alignment = align()
            c.border = border
        ws3.row_dimensions[bt_header_row].height = 24

        for ridx, row in backtest.iterrows():
            r = bt_header_row + ridx + 1
            is_win = row["total_return_pct"] > 0
            bt_vals = [
                row["rank"], row["ticker"], row["name"], row["total_score"],
                row["div_yield_pct"], row["price_1m_ago"], row["price_now"],
                row["price_change_pct"], row["cost_jpy"], row["value_now_jpy"],
                row["pnl_jpy"], row["monthly_div_jpy"],
                row["total_return_jpy"], row["total_return_pct"],
            ]
            for cidx, val in enumerate(bt_vals, start=1):
                c = ws3.cell(r, cidx, value=val)
                c.border = border
                c.alignment = align("left" if cidx == 3 else "right" if cidx >= 6 else "center")
                c.font = Font(name="Meiryo UI", size=9)
                if cidx == 14:  # 総リターン%
                    c.font = Font(name="Meiryo UI", size=9, bold=True,
                                 color=C_GREEN if is_win else C_RED)
                if ridx % 2 == 0:
                    c.fill = h_fill(C_GRAY)
            ws3.row_dimensions[r].height = 17

    # ────────────── Sheet 5: 1か月先予測 ──────────────
    ws4 = wb.create_sheet("🔮1か月先予測")
    ws4.sheet_view.showGridLines = False

    # 説明
    model_desc = [
        ("📐 予測モデルについて", ""),
        ("モデル名", "線形回帰（OLS）+ ニューロスコア補正"),
        ("予測期間", "約21営業日（1か月）"),
        ("学習データ", "過去1年間の終値（日次）"),
        ("ニューロ補正", "総合スコア差 × 2% の価格補正（高スコア銘柄を微小に上方修正）"),
        ("1σ区間", "過去残差の標準偏差（実際の価格の約68%がこの区間に収まる統計的目安）"),
        ("", ""),
        ("⚠ 重要: 本予測は統計的なトレンド延長であり、将来を保証するものではありません。", ""),
        ("  株価は多数の要因（マクロ経済、金利、企業業績、市場心理等）に影響され、", ""),
        ("  モデルの予測精度は限定的です。投資判断の唯一の根拠にしないでください。", ""),
    ]
    for ri, (k, v) in enumerate(model_desc, start=1):
        ws4.row_dimensions[ri].height = 18
        ck = ws4.cell(ri, 1, value=k)
        cv = ws4.cell(ri, 2, value=v)
        if "📐" in k:
            ck.font = Font(name="Meiryo UI", bold=True, size=12, color=C_HEADER)
            ws4.merge_cells(f"A{ri}:F{ri}")
        elif "⚠" in k or k.startswith("  "):
            ck.font = Font(name="Meiryo UI", size=9, color=C_RED, bold="⚠" in k)
            ws4.merge_cells(f"A{ri}:F{ri}")
        else:
            ck.font = Font(name="Meiryo UI", bold=True, size=10, color="374151")
            cv.font = Font(name="Meiryo UI", size=10, color="374151")
        ck.alignment = align("left")

    if not prediction.empty:
        pred_header_row = len(model_desc) + 2
        pred_cols = [
            ("スコア順位", 8), ("コード", 8), ("銘柄名", 22), ("総合スコア", 10),
            ("利回%", 8), ("現在株価", 10), ("1か月先予測", 11), ("予測変化%", 9),
            ("下限(1σ)", 10), ("上限(1σ)", 10), ("年率ﾎﾞﾗ%", 9),
            ("日次ﾄﾚﾝﾄﾞ", 10), ("モデル", 22),
        ]
        ws4.column_dimensions["A"].width = 8
        for cidx, (label, width) in enumerate(pred_cols, start=1):
            ws4.column_dimensions[get_column_letter(cidx)].width = width
            c = ws4.cell(pred_header_row, cidx, value=label)
            c.fill = h_fill(C_SUBHEAD)
            c.font = h_font()
            c.alignment = align()
            c.border = border
        ws4.row_dimensions[pred_header_row].height = 24

        for ridx, row in prediction.iterrows():
            r = pred_header_row + ridx + 1
            is_up = row["change_pct_1m"] > 0
            pred_vals = [
                row["rank"], row["ticker"], row["name"], row["total_score"],
                row["div_yield_pct"], row["price_now"],
                row["price_pred_1m"], row["change_pct_1m"],
                row["lower_1sigma"], row["upper_1sigma"],
                row.get("vol_annual_pct", ""), row["trend_slope_daily"], row["model"],
            ]
            for cidx, val in enumerate(pred_vals, start=1):
                c = ws4.cell(r, cidx, value=val)
                c.border = border
                c.alignment = align("left" if cidx in [3, len(pred_cols)] else "right" if cidx >= 6 else "center")
                c.font = Font(name="Meiryo UI", size=9)
                if cidx == 8:  # 予測変化%
                    c.font = Font(name="Meiryo UI", size=9, bold=True,
                                 color=C_GREEN if is_up else C_RED)
            ws4.row_dimensions[r].height = 17

    # ────────────── Sheet 6: 用語解説 ──────────────
    ws5 = wb.create_sheet("📖用語解説")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 22
    ws5.column_dimensions["B"].width = 70

    glossary = [
        ("用語", "説明"),
        ("ニューロファイナンス", "神経科学と金融学の融合分野。fMRI等を用いて投資判断の神経基盤を解明する。"),
        ("NAcc（側坐核）", "報酬期待に反応する脳領域。過活性化すると衝動的なリスク行動につながる[P1]。"),
        ("Insula（島皮質）", "リスクの分散（変動性）を符号化する脳領域。高活性化でパニック売りを誘発[P3]。"),
        ("損失回避係数 λ=2", "損失の神経インパクトは同額の利得の約2倍。Tom et al. (2007)より。"),
        ("ドーパミン予測誤差", "期待を超えた報酬が来るとドーパミンが発火。連続増配はこれを積み重ねる[P4]。"),
        ("", ""),
        ("安定性スコア", "配当の変動係数(CV)が小さいほど高スコア。減配歴があると事実上0に近い値になる。"),
        ("財務スコア", "自己資本比率とD/Eレシオから計算。財務健全性の代理指標。"),
        ("利回スコア", "利回り4〜6%が最高評価。8%超から下降、10%超で重罰（NAcc過活性ゾーン）。"),
        ("成長スコア", "直近3年の配当CAGR。マイナスはλ=2で急激にスコア低下。"),
        ("性向スコア", "配当性向30〜70%が理想。80%超でλ=2ペナルティ。100%超は事実上0。"),
        ("連配スコア", "連続増配年数。対数スケールで評価（初期効果が大きく、後年は逓減）。"),
        ("業種内スコア", "同業種内での相対的な利回り順位。社会比較効果の参照点として使用。"),
        ("総合スコア", "7軸のスコアを重み付き平均した値（0〜1）。1.0が理論上の最高。"),
        ("", ""),
        ("配当CV", "配当の変動係数（標準偏差÷平均）。0.10以下が安定的、0.30超は変動大。"),
        ("3年CAGR", "直近3年の年間配当複利成長率。正→増加傾向、負→減少傾向。"),
        ("D/Eレシオ", "有利子負債÷自己資本。高いほど財務リスクが高い。1.0以下が望ましい。"),
        ("", ""),
        ("引用論文", ""),
        ("[P1] Kuhnen & Knutson (2005)", "Neuron 47:763-770「金融リスクテイキングの神経基盤」"),
        ("[P2] Tom et al. (2007)", "Science 315:515-518「リスク下意思決定における損失回避の神経基盤」"),
        ("[P3] Knutson & Bossaerts (2007)", "J. Neuroscience 27:8174-8177「金融決定の神経的前兆」"),
        ("[P4] Schultz, Dayan & Montague (1997)", "Science 275:1593-1599「ドーパミン予測誤差理論」"),
        ("[P5] Frydman & Camerer (2016)", "Trends in Cognitive Sciences「金融意思決定の心理・神経科学」"),
    ]
    for ri, (k, v) in enumerate(glossary, start=1):
        ws5.row_dimensions[ri].height = 22
        ck = ws5.cell(ri, 1, value=k)
        cv = ws5.cell(ri, 2, value=v)
        if ri == 1:
            ck.fill = h_fill(C_HEADER); cv.fill = h_fill(C_HEADER)
            ck.font = h_font(); cv.font = h_font()
        elif k.startswith("[P"):
            ck.font = Font(name="Meiryo UI", size=9, color="1D4ED8", bold=True)
            cv.font = Font(name="Meiryo UI", size=9, color="374151")
        elif k == "引用論文" or k == "ニューロファイナンス":
            ck.font = Font(name="Meiryo UI", size=10, bold=True, color=C_HEADER)
            cv.font = Font(name="Meiryo UI", size=10, color="374151")
        else:
            ck.font = Font(name="Meiryo UI", size=9, color="1F2937", bold=bool(v))
            cv.font = Font(name="Meiryo UI", size=9, color="374151")
        ck.alignment = align("left"); cv.alignment = align("left", wrap=True)
        ck.border = border; cv.border = border
        if ri % 2 == 0:
            ck.fill = h_fill(C_GRAY); cv.fill = h_fill(C_GRAY)

    wb.save(output_path)
    log.info(f"✅ レポート保存完了: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("JPX高配当株 ニューロファイナンス・スコアリング 分析開始")
    print("=" * 60)

    # 1. JPX銘柄リスト取得
    tickers_df = fetch_jpx_ticker_list()

    # 2. yfinanceでデータ取得（最大300銘柄）
    raw_df = fetch_stock_data(tickers_df, max_tickers=300)

    if raw_df.empty:
        log.error("データ取得に失敗しました。ネットワーク接続を確認してください。")
        return

    # 3. スコアリング
    scored_df = compute_scores(raw_df)

    print(f"\n📊 上位10銘柄:")
    print(scored_df[["rank","ticker","name","div_yield_pct","total_score",
                     "s_stability","s_health","s_yield"]].head(10).to_string(index=False))

    # 4. バックテスト（上位20銘柄）
    bt_df = run_backtest(scored_df, top_n=20)

    # 5. 1か月先予測（上位20銘柄）
    pred_df = predict_1month(scored_df, top_n=20)

    # 6. Excelレポート出力
    write_excel_report(scored_df, bt_df, pred_df, OUTPUT_PATH)

    print("\n" + "=" * 60)
    print(f"✅ 完了! レポートを開く:")
    print(f"   {OUTPUT_PATH}")
    print("=" * 60)


if __name__ == "__main__":
    main()
