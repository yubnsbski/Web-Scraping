# JPX NeuroScore Excel v3 - 実績・クラスタ統合レポート
import sys
_log = open(r"C:\Users\ynobe\Desktop\excel_v3_log.txt", "w", encoding="utf-8")
sys.stdout = _log
sys.stderr = _log

import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

SRC = r"C:\Users\ynobe\Desktop\jpx_ml_results.json"
V2  = r"C:\Users\ynobe\Desktop\jpx_ml_v2_results.json"
OUT = r"C:\Users\ynobe\Desktop\JPX_NeuroScore_Report_v3.xlsx"

print("=== JPX Excel v3 生成 ===")
with open(SRC, encoding='utf-8') as f:
    data = json.load(f)
stocks = data['stocks']

v2_data = {}
try:
    with open(V2, encoding='utf-8') as f:
        v2_data = json.load(f)
    print("v2結果読み込み完了")
except Exception as e:
    print(f"v2結果なし（スキップ）: {e}")

# --- スタイル定義 ---
def hdr(ws, row, col, val, bg='1F3864', fg='FFFFFF', bold=True, sz=11, align='center', wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name='Arial', bold=bold, color=fg, size=sz)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    return c

def cell(ws, row, col, val, bold=False, sz=10, align='center', num_fmt=None, color=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name='Arial', bold=bold, size=sz, color=color or '000000')
    c.alignment = Alignment(horizontal=align, vertical='center')
    if num_fmt:
        c.number_format = num_fmt
    return c

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

CLUSTER_COLORS = {0: 'DBEEFF', 1: 'D5F5E3', 2: 'FEF9E7', 3: 'FDEDEC'}
CLUSTER_NAMES  = {0: 'C0: 財務優良型', 1: 'C1: 中財務成長型', 2: 'C2: 高成長型', 3: 'C3: 高利回型'}

wb = Workbook()

# ===========================
# Sheet 1: ダッシュボード
# ===========================
ws1 = wb.active
ws1.title = 'ダッシュボード'
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 2
for col, w in enumerate([18,14,14,14,14,14,14,14], start=2):
    ws1.column_dimensions[get_column_letter(col)].width = w

hdr(ws1, 1, 2, 'JPX NeuroFinance スコアレポート v3', bg='1F3864', sz=16, bold=True)
ws1.merge_cells('B1:I1')
ws1.row_dimensions[1].height = 36

hdr(ws1, 2, 2, f'生成日: {datetime.now().strftime("%Y-%m-%d %H:%M")}  対象: {len(stocks)}銘柄', bg='2E4374', sz=10, bold=False)
ws1.merge_cells('B2:I2')

# モデル精度サマリー
hdr(ws1, 4, 2, '教師あり vs 教師なし 比較', bg='2E4374', sz=12)
ws1.merge_cells('B4:I4')

sup = data['supervised']
unsup = data['unsupervised']
comp = data['comparison']

headers = ['指標', '教師あり(RF)', '教師あり(Ridge)', '教師あり(OLS)', '教師なし(K-means)', '判定']
for ci, h in enumerate(headers, start=2):
    hdr(ws1, 5, ci, h, bg='34495E', sz=10)

rows_data = [
    ['CV R²', sup['models']['RandomForest']['cv_r2_mean'], sup['models']['Ridge（正則化線形）']['cv_r2_mean'],
     sup['models']['OLS（線形回帰）']['cv_r2_mean'], unsup['r2_vs_actual'], '教師なし優位'],
    ['MAE (%)', sup['models']['RandomForest']['mae_pct'], sup['models']['Ridge（正則化線形）']['mae_pct'],
     sup['models']['OLS（線形回帰）']['mae_pct'], unsup['mae_pct'], 'RF最小'],
    ['訓練R²', sup['models']['RandomForest']['train_r2'], sup['models']['Ridge（正則化線形）']['train_r2'],
     sup['models']['OLS（線形回帰）']['train_r2'], '-', '過学習注意'],
]

for ri, row in enumerate(rows_data, start=6):
    for ci, val in enumerate(row, start=2):
        c = ws1.cell(row=ri, column=ci, value=val)
        c.font = Font(name='Arial', size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()
        if isinstance(val, float):
            c.number_format = '0.000'
            if val < 0:
                c.font = Font(name='Arial', size=10, color='C0392B')
            elif val > 0.1:
                c.font = Font(name='Arial', size=10, color='1A5276')

# 外れ値情報
hdr(ws1, 10, 2, '外れ値情報', bg='2E4374', sz=12)
ws1.merge_cells('B10:I10')
if v2_data:
    info_rows = [
        ['IQR範囲', f"{v2_data['outlier_bounds']['lo']:.1f}% ～ {v2_data['outlier_bounds']['hi']:.1f}%"],
        ['外れ値銘柄数', f"{v2_data['n_outliers']}銘柄"],
        ['外れ値除外後 CV R²', f"{v2_data['comparison_summary']['supervised_cv_r2_after']:.4f}"],
        ['改善幅', f"Δ{v2_data['comparison_summary']['supervised_cv_r2_after'] - v2_data['comparison_summary']['supervised_cv_r2_before']:+.4f}"],
    ]
    for ri, (k, v) in enumerate(info_rows, start=11):
        cell(ws1, ri, 2, k, bold=True, align='left')
        cell(ws1, ri, 3, v, align='left')
        ws1.merge_cells(f'C{ri}:I{ri}')

# クラスタサマリー
hdr(ws1, 16, 2, 'クラスタ別パフォーマンス', bg='2E4374', sz=12)
ws1.merge_cells('B16:I16')
cl_headers = ['クラスタ', '銘柄数', '平均1年リターン', '標準偏差', 'NeuroScore平均', '代表銘柄(上位3)']
for ci, h in enumerate(cl_headers, start=2):
    hdr(ws1, 17, ci, h, bg='34495E', sz=10)
for ci_key in range(4):
    cs = unsup['cluster_stats'].get(f'C{ci_key}', {})
    ri = 18 + ci_key
    fill_color = CLUSTER_COLORS[ci_key]
    values = [
        CLUSTER_NAMES[ci_key],
        cs.get('count', '-'),
        f"{cs.get('avg_ret_1y', 0):.1f}%",
        f"±{cs.get('std_ret_1y', 0):.1f}%",
        f"{cs.get('avg_neuro_score', 0):.4f}",
        '、'.join(cs.get('top5_names', [])[:3]),
    ]
    for ci, v in enumerate(values, start=2):
        c = ws1.cell(row=ri, column=ci, value=v)
        c.font = Font(name='Arial', size=10)
        c.fill = PatternFill('solid', fgColor=fill_color)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
ws1.row_dimensions[22].height = 15

# ===========================
# Sheet 2: スコアランキング（全127銘柄）
# ===========================
ws2 = wb.create_sheet('スコアランキング')
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = 'A3'

col_defs = [
    ('順位', 5), ('コード', 7), ('銘柄名', 16), ('セクター', 14),
    ('クラスタ', 10), ('NeuroScore', 11), ('安定性', 9), ('財務', 9),
    ('利回得点', 9), ('成長性', 9), ('性向', 8), ('連配', 8), ('業種内', 8),
    ('RMS', 8), ('1年実績%', 10), ('教師あり予測%', 11), ('実配当%', 10), ('現在株価', 10),
]
for ci, (name, width) in enumerate(col_defs, start=1):
    ws2.column_dimensions[get_column_letter(ci)].width = width
    hdr(ws2, 2, ci, name, bg='1F3864', sz=9)
ws2.row_dimensions[2].height = 30

# タイトル
hdr(ws2, 1, 1, 'JPX NeuroScore ランキング（127銘柄 | 実績・クラスタ統合）', bg='1F3864', sz=13)
ws2.merge_cells(f'A1:{get_column_letter(len(col_defs))}1')

# データ行
sorted_stocks = sorted(stocks, key=lambda x: x['total_score'], reverse=True)
for ri, s in enumerate(sorted_stocks, start=3):
    cl = s.get('cluster')
    fill = PatternFill('solid', fgColor=CLUSTER_COLORS.get(cl, 'FFFFFF'))
    values = [
        ri - 2,  # rank
        s['ticker'],
        s['name'],
        s['sector'],
        f"C{cl}" if cl is not None else '-',
        s['total_score'],
        s['scores']['s_stability'],
        s['scores']['s_health'],
        s['scores']['s_yield'],
        s['scores']['s_momentum'],
        s['scores']['s_payout'],
        s['scores']['s_streak'],
        s['scores']['s_sector'],
        s.get('rms'),
        s.get('actual_ret_1y'),
        s.get('pred_ret_supervised'),
        s.get('actual_yield_pct'),
        s.get('current_price'),
    ]
    for ci, val in enumerate(values, start=1):
        c = ws2.cell(row=ri, column=ci, value=val)
        c.font = Font(name='Arial', size=9)
        c.fill = fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()
        if ci in [6,7,8,9,10,11,12,13,14] and isinstance(val, float):
            c.number_format = '0.000'
        if ci in [15,16]:
            if isinstance(val, float):
                c.number_format = '0.0'
                if val < 0:
                    c.font = Font(name='Arial', size=9, color='C0392B')
                elif val > 100:
                    c.font = Font(name='Arial', size=9, color='1A5276', bold=True)
        if ci == 17 and isinstance(val, float):
            c.number_format = '0.00'
        if ci == 18 and isinstance(val, float):
            c.number_format = '#,##0'

# 条件付きフォーマット: NeuroScore列
ws2.conditional_formatting.add(
    f'F3:F{2+len(stocks)}',
    ColorScaleRule(start_type='min', start_color='FFF9C4',
                   mid_type='percentile', mid_value=50, mid_color='FFEB3B',
                   end_type='max', end_color='F57F17')
)
# 1年実績列
ws2.conditional_formatting.add(
    f'O3:O{2+len(stocks)}',
    ColorScaleRule(start_type='min', start_color='FFCDD2',
                   mid_type='num', mid_value=0, mid_color='FFFFFF',
                   end_type='max', end_color='C8E6C9')
)

# ===========================
# Sheet 3: ML分析
# ===========================
ws3 = wb.create_sheet('ML分析')
ws3.sheet_view.showGridLines = False
for col, w in enumerate([2,20,14,14,14,14,14], start=1):
    ws3.column_dimensions[get_column_letter(col)].width = w

hdr(ws3, 1, 2, '教師あり vs 教師なし 詳細比較', bg='1F3864', sz=14)
ws3.merge_cells('B1:G1')

# モデル比較テーブル
hdr(ws3, 3, 2, 'モデル', bg='34495E'); hdr(ws3, 3, 3, 'CV R²', bg='34495E')
hdr(ws3, 3, 4, 'CV std', bg='34495E'); hdr(ws3, 3, 5, '訓練R²', bg='34495E')
hdr(ws3, 3, 6, 'MAE(%)', bg='34495E'); hdr(ws3, 3, 7, '評価', bg='34495E')

model_rows = [
    ('OLS（線形回帰）', sup['models']['OLS（線形回帰）']['cv_r2_mean'],
     sup['models']['OLS（線形回帰）']['cv_r2_std'], sup['models']['OLS（線形回帰）']['train_r2'],
     sup['models']['OLS（線形回帰）']['mae_pct'], '参考値'),
    ('Ridge（正則化）', sup['models']['Ridge（正則化線形）']['cv_r2_mean'],
     sup['models']['Ridge（正則化線形）']['cv_r2_std'], sup['models']['Ridge（正則化線形）']['train_r2'],
     sup['models']['Ridge（正則化線形）']['mae_pct'], '参考値'),
    ('RandomForest', sup['models']['RandomForest']['cv_r2_mean'],
     sup['models']['RandomForest']['cv_r2_std'], sup['models']['RandomForest']['train_r2'],
     sup['models']['RandomForest']['mae_pct'], '最良（過学習）'),
    ('GradientBoosting', sup['models']['GradientBoosting']['cv_r2_mean'],
     sup['models']['GradientBoosting']['cv_r2_std'], sup['models']['GradientBoosting']['train_r2'],
     sup['models']['GradientBoosting']['mae_pct'], '訓練過適合'),
    ('K-means (k=4)', '-', '-', unsup['r2_vs_actual'], unsup['mae_pct'], '教師なし最良'),
]
for ri, row in enumerate(model_rows, start=4):
    for ci, val in enumerate(row, start=2):
        c = ws3.cell(row=ri, column=ci, value=val)
        c.font = Font(name='Arial', size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()
        if isinstance(val, float) and ci in [3,4,5]:
            c.number_format = '0.0000'
            if val < 0:
                c.font = Font(name='Arial', size=10, color='C0392B')
        if isinstance(val, float) and ci == 6:
            c.number_format = '0.0'

# 特徴量重要度
hdr(ws3, 10, 2, '特徴量重要度 (RandomForest)', bg='2E4374', sz=12)
ws3.merge_cells('B10:G10')
hdr(ws3, 11, 2, '軸名', bg='34495E'); hdr(ws3, 11, 3, 'ラベル', bg='34495E')
hdr(ws3, 11, 4, 'RF重要度', bg='34495E'); hdr(ws3, 11, 5, 'OLS係数比', bg='34495E')
hdr(ws3, 11, 6, '設計ウエイト', bg='34495E'); hdr(ws3, 11, 7, '解釈', bg='34495E')

rf_fi   = sup['models']['RandomForest']['feature_importance']
ols_fi  = sup['models']['OLS（線形回帰）']['feature_importance']
weights = dict(zip(data['score_keys'], data['weights']))
interpretations = {
    's_stability':  '配当の安定性（最重要）',
    's_health':     '財務健全性・損失回避',
    's_yield':      '利回り最適帯 2-8%',
    's_momentum':   '配当成長率・ドーパミン効果',
    's_payout':     '配当性向 20-60%',
    's_streak':     '連続増配年数',
    's_sector':     'セクター内相対順位',
}
for ri, key in enumerate(data['score_keys'], start=12):
    label = data['score_labels'][list(data['score_keys']).index(key)]
    vals = [key, label, rf_fi[key], ols_fi[key], weights[key], interpretations[key]]
    for ci, val in enumerate(vals, start=2):
        c = ws3.cell(row=ri, column=ci, value=val)
        c.font = Font(name='Arial', size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()
        if isinstance(val, float) and ci in [4,5,6]:
            c.number_format = '0.0000'

# v2外れ値除外比較
if v2_data:
    hdr(ws3, 21, 2, '外れ値除外後の精度改善', bg='2E4374', sz=12)
    ws3.merge_cells('B21:G21')
    v2_comp = v2_data.get('comparison_summary', {})
    hdr(ws3, 22, 2, '指標', bg='34495E'); hdr(ws3, 22, 3, '全データ(n=127)', bg='34495E')
    hdr(ws3, 22, 4, '外れ値除外後', bg='34495E'); hdr(ws3, 22, 5, '改善幅', bg='34495E')
    hdr(ws3, 22, 6, '外れ値銘柄数', bg='34495E')
    ws3.merge_cells('F22:G22')
    comp_rows = [
        ['最良モデル CV R²',
         v2_comp.get('supervised_cv_r2_before', '-'),
         v2_comp.get('supervised_cv_r2_after', '-'),
         round((v2_comp.get('supervised_cv_r2_after', 0) - v2_comp.get('supervised_cv_r2_before', 0)), 4),
         f"{v2_data.get('n_outliers', '?')}銘柄"],
        ['K-means R²',
         v2_comp.get('unsupervised_r2_before', '-'),
         v2_comp.get('unsupervised_r2_after', '-'),
         round((v2_comp.get('unsupervised_r2_after', 0) - v2_comp.get('unsupervised_r2_before', 0)), 4),
         '-'],
    ]
    for ri, row in enumerate(comp_rows, start=23):
        for ci, val in enumerate(row, start=2):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.font = Font(name='Arial', size=10)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin_border()
            if isinstance(val, float):
                c.number_format = '0.0000'
                if ci == 5 and val > 0:
                    c.font = Font(name='Arial', size=10, color='1A5276', bold=True)

# ===========================
# Sheet 4: 論文解説
# ===========================
ws4 = wb.create_sheet('論文解説')
ws4.sheet_view.showGridLines = False
ws4.column_dimensions['A'].width = 2
ws4.column_dimensions['B'].width = 14
ws4.column_dimensions['C'].width = 25
ws4.column_dimensions['D'].width = 50
ws4.column_dimensions['E'].width = 12

hdr(ws4, 1, 2, 'NeuroFinance モデル論文根拠', bg='1F3864', sz=14)
ws4.merge_cells('B1:E1')

papers = data['selection_rationale']['論文根拠']
hdr(ws4, 3, 2, 'ID', bg='34495E'); hdr(ws4, 3, 3, '論文', bg='34495E')
hdr(ws4, 3, 4, '対応スコア軸', bg='34495E'); hdr(ws4, 3, 5, '重み', bg='34495E')

weights_by_axis = {
    's_health': '20%', 's_payout': '10%', 's_yield': '20%',
    's_momentum': '15%', 's_stability': '25%'
}
for ri, p in enumerate(papers, start=4):
    axis_key = p['axis'].split('（')[0].strip()
    vals = [p['id'], p['ref'], p['axis'], weights_by_axis.get(axis_key, '-')]
    for ci, val in enumerate(vals, start=2):
        c = ws4.cell(row=ri, column=ci, value=val)
        c.font = Font(name='Arial', size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws4.row_dimensions[ri].height = 30

# メカニズム詳細
hdr(ws4, 10, 2, 'モデルメカニズム詳細', bg='2E4374', sz=12)
ws4.merge_cells('B10:E10')
for ri, p in enumerate(papers, start=11):
    c1 = ws4.cell(row=ri, column=2, value=p['id'])
    c1.font = Font(name='Arial', size=10, bold=True)
    c1.alignment = Alignment(horizontal='center', vertical='center')
    c2 = ws4.cell(row=ri, column=3, value=p['mechanism'])
    c2.font = Font(name='Arial', size=10)
    c2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws4.merge_cells(f'C{ri}:E{ri}')
    ws4.row_dimensions[ri].height = 30

# スクリーニング条件
hdr(ws4, 17, 2, 'スクリーニング条件', bg='2E4374', sz=12)
ws4.merge_cells('B17:E17')
for ri, cond in enumerate(data['selection_rationale']['スクリーニング条件'], start=18):
    c = ws4.cell(row=ri, column=2, value=f'• {cond}')
    c.font = Font(name='Arial', size=10)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws4.merge_cells(f'B{ri}:E{ri}')

# 注意事項
ws4.cell(row=24, column=2, value='【データ品質注意】yfinance利回り値(121/127銘柄)はDPS(円)の場合あり。実配当利回りは配当履歴/最新株価で独自計算済み。').font = Font(name='Arial', size=9, italic=True, color='7F8C8D')
ws4.merge_cells('B24:E24')

# 保存
wb.save(OUT)
print(f"完了: {OUT}")
_log.flush()
_log.close()
