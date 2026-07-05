"""
JPX_NeuroScore_Report.xlsx 改善スクリプト
"""
import sys
_log = open(r"C:\Users\ynobe\Desktop\jpx_err.txt", "w", encoding="utf-8")
sys.stdout = _log
sys.stderr = _log
import json
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from datetime import datetime
import os

SRC = r"C:\Users\ynobe\Desktop\JPX_NeuroScore_Report.xlsx"
DST = r"C:\Users\ynobe\Desktop\JPX_NeuroScore_Report_v2.xlsx"
JSON_OUT = r"C:\Users\ynobe\Desktop\jpx_data.json"

# ── 1. データ読み込み ──────────────────────────────────────────
print("読み込み中...")
xl = pd.ExcelFile(SRC)
print("シート:", xl.sheet_names)

# 全銘柄ランキングシートを探す
rank_sheet = None
for s in xl.sheet_names:
    if "スコア" in s or "ランキング" in s or "Rank" in s.lower():
        rank_sheet = s
        break
if rank_sheet is None:
    rank_sheet = xl.sheet_names[1] if len(xl.sheet_names) > 1 else xl.sheet_names[0]

df_raw = pd.read_excel(SRC, sheet_name=rank_sheet, header=0)
print(f"列: {list(df_raw.columns)}")
print(f"行数: {len(df_raw)}")

# バックテストシート
bt_sheet = None
for s in xl.sheet_names:
    if "バック" in s or "back" in s.lower():
        bt_sheet = s
        break

df_bt = pd.DataFrame()
if bt_sheet:
    try:
        df_bt = pd.read_excel(SRC, sheet_name=bt_sheet, header=None)
    except:
        pass

# ── 2. データクレンジング ──────────────────────────────────────
# 列名直接マッピング（実際の列名に基づく）
col_map = {
    '順位': 'rank', 'コード': 'ticker', '銘柄名': 'name', 'セクター': 'sector',
    '利回り%': 'yield_pct', '連配年': 'streak', '性向%': 'payout_pct', '自己資本%': 'equity_pct',
    '安定性': 's_stability', '財務': 's_health', '利回得点': 's_yield', '成長性': 's_momentum',
    '性向得点': 's_payout', '連配得点': 's_streak', '業種内': 's_sector', '総合スコア': 'total_score',
    '減配歴': 'has_cut'
}
df = df_raw.rename(columns=col_map)

# 数値変換
for c in ["yield_pct","payout_pct","equity_pct","total_score","s_stability","s_health","s_yield","s_momentum","s_payout","s_streak","s_sector","streak"]:
    if c in df.columns:
        df[c] = pd.to_numeric(df[c], errors="coerce")

# 利回りクレンジング: yfinanceがDPS(円)を利回%として返すケースを除外
# 利回り > 20% は異常値とする（実際には存在しない）
if "yield_pct" in df.columns:
    # 正常値フラグ
    df["yield_valid"] = df["yield_pct"].between(0.01, 20.0, inclusive="both")
    n_invalid = (~df["yield_valid"]).sum()
    print(f"利回り異常値: {n_invalid}銘柄（除外）")
    # スコアはそのまま使う（利回り異常でも他の指標は正常な可能性）

# 利回り表示用（異常値はNaN）
if "yield_pct" in df.columns:
    df["yield_display"] = df["yield_pct"].where(df["yield_valid"])

# total_scoreが有効な行のみ
if "total_score" in df.columns:
    df = df.dropna(subset=["total_score"])
    df = df.sort_values("total_score", ascending=False).reset_index(drop=True)
    df["rank_new"] = range(1, len(df)+1)

print(f"有効銘柄数: {len(df)}")

# ── 3. JSON出力（HTML用）────────────────────────────────────────
score_cols = [c for c in ["s_stability","s_health","s_yield","s_momentum","s_payout","s_streak","s_sector"] if c in df.columns]
records = []
for _, row in df.iterrows():
    r = {
        "rank": int(row.get("rank_new", row.get("rank", 0))),
        "ticker": str(row.get("ticker", "")),
        "name": str(row.get("name", "")),
        "sector": str(row.get("sector", "")),
        "total_score": round(float(row.get("total_score", 0)), 4),
        "yield_pct": round(float(row["yield_display"]), 2) if "yield_display" in df.columns and pd.notna(row.get("yield_display")) else None,
        "streak": int(row.get("streak", 0)) if pd.notna(row.get("streak", float("nan"))) else 0,
        "payout_pct": round(float(row["payout_pct"]), 1) if "payout_pct" in df.columns and pd.notna(row.get("payout_pct")) else None,
        "equity_pct": round(float(row["equity_pct"]), 1) if "equity_pct" in df.columns and pd.notna(row.get("equity_pct")) else None,
        "has_cut": str(row.get("has_cut", "")) if "has_cut" in df.columns else "",
        "scores": {c: round(float(row[c]), 4) if pd.notna(row.get(c)) else 0 for c in score_cols},
    }
    records.append(r)

with open(JSON_OUT, "w", encoding="utf-8") as f:
    json.dump({"generated": datetime.now().isoformat(), "count": len(records), "stocks": records}, f, ensure_ascii=False, indent=2)
print(f"JSON出力: {JSON_OUT}")

# ── 4. 新Excelファイル構築 ────────────────────────────────────
from openpyxl import Workbook
wb = Workbook()
wb.remove(wb.active)

# スタイル定義
C_NAVY = "1E3A5F"
C_GOLD = "D4A843"
C_LIGHT = "EEF4FB"
C_WHITE = "FFFFFF"
C_GRAY = "F3F4F6"
C_GREEN = "16A34A"
C_RED = "DC2626"
C_BLUE = "2563EB"
border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
thin_bottom = Border(bottom=Side(style="medium", color=C_NAVY))

def hfont(bold=True, size=10, color="FFFFFF", name="Meiryo UI"):
    return Font(name=name, bold=bold, size=size, color=color)
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)
def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Sheet 1: ダッシュボード ────────────────────────────────────
ws0 = wb.create_sheet("📊ダッシュボード")
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 2
ws0.column_dimensions["B"].width = 14
ws0.column_dimensions["C"].width = 26
ws0.column_dimensions["D"].width = 12
ws0.column_dimensions["E"].width = 12
ws0.column_dimensions["F"].width = 12
ws0.column_dimensions["G"].width = 14

# タイトル
ws0.merge_cells("B1:G1")
c = ws0["B1"]
c.value = "JPX ニューロファイナンス・スコアリング　ダッシュボード"
c.font = Font(name="Meiryo UI", bold=True, size=16, color=C_NAVY)
c.alignment = align("left")
ws0.row_dimensions[1].height = 36

ws0.merge_cells("B2:G2")
c = ws0["B2"]
c.value = f"分析日: {datetime.now().strftime('%Y年%m月%d日')}　　有効銘柄数: {len(df)}銘柄"
c.font = Font(name="Meiryo UI", size=10, color="6B7280")
c.alignment = align("left")
ws0.row_dimensions[2].height = 22

# KPIカード（行4-8）
ws0.row_dimensions[3].height = 12
top5 = df.head(5)
valid_yield = df[df.get("yield_valid", pd.Series([True]*len(df)))] if "yield_valid" in df.columns else df
avg_yield = valid_yield["yield_display"].mean() if "yield_display" in df.columns and not valid_yield.empty else 0
avg_score = df["total_score"].mean() if "total_score" in df.columns else 0
streak_valid = df[df["streak"] > 0]["streak"] if "streak" in df.columns else pd.Series([])
avg_streak = streak_valid.mean() if not streak_valid.empty else 0

kpis = [
    ("分析銘柄数", f"{len(df)}銘柄", C_NAVY),
    ("平均ニューロスコア", f"{avg_score:.3f}", C_BLUE),
    ("平均配当利回り*", f"{avg_yield:.1f}%" if avg_yield else "---", C_GREEN),
    ("平均連続増配", f"{avg_streak:.0f}年" if avg_streak else "---", "6D28D9"),
]
for i, (label, val, color) in enumerate(kpis):
    col = chr(ord("B") + i)
    ws0.merge_cells(f"{col}4:{col}5")
    ws0.merge_cells(f"{col}6:{col}7")
    c_lbl = ws0[f"{col}4"]
    c_val = ws0[f"{col}6"]
    c_lbl.value = label
    c_lbl.font = Font(name="Meiryo UI", size=9, bold=True, color=color)
    c_lbl.alignment = align()
    c_lbl.fill = fill("F8FAFC")
    c_lbl.border = Border(top=Side(style="medium", color=color), left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"))
    c_val.value = val
    c_val.font = Font(name="Meiryo UI", size=20, bold=True, color=color)
    c_val.alignment = align()
    c_val.border = Border(bottom=Side(style="medium", color=color), left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"))
    for r in [4,5,6,7]:
        ws0.row_dimensions[r].height = 20

ws0.row_dimensions[8].height = 14

# 上位10銘柄テーブル
ws0.merge_cells("B9:G9")
c = ws0["B9"]
c.value = "▼ ニューロスコア 上位10銘柄"
c.font = Font(name="Meiryo UI", bold=True, size=11, color=C_WHITE)
c.alignment = align("left")
c.fill = fill(C_NAVY)
ws0.row_dimensions[9].height = 24

top_headers = ["順位", "コード", "銘柄名", "利回り%*", "総合スコア", "財務健全性"]
top_widths = [6, 8, 26, 10, 12, 12]
for ci, (h, w) in enumerate(zip(top_headers, top_widths), start=2):
    cl = get_column_letter(ci)
    c = ws0.cell(10, ci, value=h)
    c.font = Font(name="Meiryo UI", bold=True, size=9, color=C_WHITE)
    c.fill = fill("374151")
    c.alignment = align()
    c.border = border
ws0.row_dimensions[10].height = 20

for ri, (_, row) in enumerate(df.head(10).iterrows(), start=11):
    bg = "FEF3C7" if ri <= 13 else ("DBEAFE" if ri <= 16 else C_WHITE)
    row_vals = [
        row.get("rank_new", ri - 10),
        row.get("ticker", ""),
        row.get("name", ""),
        f"{row['yield_display']:.1f}%" if "yield_display" in df.columns and pd.notna(row.get("yield_display")) else "---",
        round(float(row.get("total_score", 0)), 4),
        f"{row.get('s_health', 0):.3f}" if pd.notna(row.get("s_health", float("nan"))) else "---",
    ]
    for ci, val in enumerate(row_vals, start=2):
        c = ws0.cell(ri, ci, value=val)
        c.font = Font(name="Meiryo UI", size=9, bold=(ri <= 13))
        c.fill = fill(bg)
        c.alignment = align("left" if ci == 4 else "center")
        c.border = border
    ws0.row_dimensions[ri].height = 18

ws0.merge_cells("B21:G21")
ws0["B21"].value = "* 利回り20%超のデータはyfinanceの取得誤差（DPS円値が混入）のため除外。実際の利回りはYahoo Finance等で確認してください。"
ws0["B21"].font = Font(name="Meiryo UI", size=8, color="6B7280", italic=True)
ws0["B21"].alignment = align("left")

# ── Sheet 2: スコアランキング（フルリスト）────────────────────
ws1 = wb.create_sheet("🏆スコアランキング")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 2

rank_cols = [
    ("順位", 7), ("コード", 9), ("銘柄名", 22), ("セクター", 18),
    ("利回り%", 9), ("連配年", 8), ("性向%", 8), ("自己資本%", 9),
    ("安定性", 9), ("財務", 9), ("利回得点", 9), ("成長性", 9),
    ("性向得点", 9), ("連配得点", 9), ("業種内", 8), ("総合スコア", 11),
    ("減配歴", 8),
]
ws1.merge_cells(f"B1:{get_column_letter(1+len(rank_cols))}1")
c = ws1["B1"]
c.value = "JPX 高配当銘柄 ニューロスコア ランキング（全銘柄）"
c.font = Font(name="Meiryo UI", bold=True, size=14, color=C_NAVY)
c.alignment = align("left")
ws1.row_dimensions[1].height = 32

for ci, (label, width) in enumerate(rank_cols, start=2):
    ws1.column_dimensions[get_column_letter(ci)].width = width
    c = ws1.cell(2, ci, value=label)
    c.fill = fill(C_NAVY)
    c.font = Font(name="Meiryo UI", bold=True, size=9, color=C_WHITE)
    c.alignment = align()
    c.border = border
ws1.row_dimensions[2].height = 22

col_keys = ["rank_new","ticker","name","sector","yield_display","streak","payout_pct","equity_pct",
            "s_stability","s_health","s_yield","s_momentum","s_payout","s_streak","s_sector","total_score","has_cut"]

for ridx, (_, row) in enumerate(df.iterrows(), start=3):
    rank = int(row.get("rank_new", ridx - 2))
    is_top3 = rank <= 3
    is_top10 = rank <= 10
    bg = "FEF3C7" if is_top3 else ("F0F9FF" if is_top10 else (C_GRAY if (ridx % 2 == 0) else C_WHITE))

    for ci, key in enumerate(col_keys, start=2):
        val = row.get(key)
        # 表示変換
        if key == "yield_display":
            val = f"{float(val):.1f}%" if pd.notna(val) else "---"
        elif key == "payout_pct":
            val = round(float(val), 1) if pd.notna(val) else "---"
        elif key == "equity_pct":
            val = round(float(val), 1) if pd.notna(val) else "---"
        elif key in ["s_stability","s_health","s_yield","s_momentum","s_payout","s_streak","s_sector","total_score"]:
            val = round(float(val), 4) if pd.notna(val) else "---"
        elif key == "has_cut":
            val = str(val) if val and str(val) not in ["nan","None",""] else "なし"
        elif key == "streak":
            val = int(val) if pd.notna(val) and val != "" else 0

        c = ws1.cell(ridx, ci, value=val)
        c.fill = fill(bg)
        c.border = border
        c.alignment = align("left" if ci == 4 else "center")
        c.font = Font(name="Meiryo UI", size=9, bold=is_top3,
                      color=C_NAVY if is_top3 else ("374151" if is_top10 else "6B7280"))

        # 総合スコア色分け
        if key == "total_score" and isinstance(val, float):
            if val >= 0.55:
                c.fill = fill("DCFCE7"); c.font = Font(name="Meiryo UI", size=9, bold=True, color="166534")
            elif val >= 0.50:
                c.fill = fill("DBEAFE"); c.font = Font(name="Meiryo UI", size=9, bold=True, color="1E40AF")

    ws1.row_dimensions[ridx].height = 16

# 条件付き書式（総合スコア列）
score_col_letter = get_column_letter(2 + col_keys.index("total_score"))
last_row = 2 + len(df)
ws1.conditional_formatting.add(
    f"{score_col_letter}3:{score_col_letter}{last_row}",
    ColorScaleRule(start_type="min", start_color="FFFFFF",
                   mid_type="percentile", mid_value=50, mid_color="DBEAFE",
                   end_type="max", end_color="DCFCE7")
)

# ── Sheet 3: グラフ ───────────────────────────────────────────
ws_chart = wb.create_sheet("📈グラフ")
ws_chart.sheet_view.showGridLines = False
ws_chart.column_dimensions["A"].width = 2

ws_chart.merge_cells("B1:R1")
c = ws_chart["B1"]
c.value = "ニューロスコア グラフ分析"
c.font = Font(name="Meiryo UI", bold=True, size=14, color=C_NAVY)
c.alignment = align("left")
ws_chart.row_dimensions[1].height = 30

# グラフ用データ（上位20銘柄）
top20 = df.head(20)
ws_chart["B3"].value = "銘柄名"; ws_chart["B3"].font = Font(bold=True, size=9, color=C_WHITE)
ws_chart["B3"].fill = fill(C_NAVY); ws_chart["B3"].alignment = align()
score_sub_cols = ["総合スコア","安定性","財務","利回得点","成長性","性向得点","連配得点"]
score_keys2 = ["total_score","s_stability","s_health","s_yield","s_momentum","s_payout","s_streak"]
for ci_off, label in enumerate(score_sub_cols, start=3):
    c = ws_chart.cell(3, ci_off, value=label)
    c.font = Font(bold=True, size=8, color=C_WHITE)
    c.fill = fill(C_NAVY)
    c.alignment = align()
    ws_chart.column_dimensions[get_column_letter(ci_off)].width = 10

ws_chart.column_dimensions["B"].width = 22
ws_chart.row_dimensions[3].height = 20

for ri, (_, row) in enumerate(top20.iterrows(), start=4):
    name = str(row.get("name", ""))[:10]
    ws_chart.cell(ri, 2, value=name)
    ws_chart.cell(ri, 2).font = Font(name="Meiryo UI", size=8)
    ws_chart.cell(ri, 2).alignment = align("left")
    for ci_off, key in enumerate(score_keys2, start=3):
        val = row.get(key)
        v = round(float(val), 4) if pd.notna(val) else 0
        c = ws_chart.cell(ri, ci_off, value=v)
        c.font = Font(name="Meiryo UI", size=8)
        c.alignment = align()
        if ci_off == 3:  # 総合スコア
            c.font = Font(name="Meiryo UI", size=8, bold=True, color=C_BLUE)
    ws_chart.row_dimensions[ri].height = 15

# 棒グラフ：上位20銘柄の総合スコア
bar = BarChart()
bar.type = "bar"  # 横棒
bar.title = "ニューロスコア 上位20銘柄"
bar.y_axis.title = "銘柄"
bar.x_axis.title = "総合スコア"
bar.style = 10
bar.grouping = "clustered"
bar.width = 20
bar.height = 14

data_ref = Reference(ws_chart, min_col=3, max_col=3, min_row=3, max_row=3+len(top20))
cats_ref = Reference(ws_chart, min_col=2, min_row=4, max_row=3+len(top20))
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.series[0].graphicalProperties.solidFill = "2563EB"
ws_chart.add_chart(bar, "J3")

# スタック棒グラフ：7軸スコア内訳（上位10）
bar2 = BarChart()
bar2.type = "bar"
bar2.title = "スコア内訳（7軸）上位10銘柄"
bar2.style = 10
bar2.grouping = "stacked"
bar2.overlap = 100
bar2.width = 20
bar2.height = 14

colors_7 = ["1E3A5F","2563EB","16A34A","D97706","7C3AED","DC2626","0891B2"]
score_labels = ["安定性","財務","利回","成長","性向","連配","業種"]
for i, (key, color, lbl) in enumerate(zip(score_keys2[1:], colors_7, score_labels)):
    col_idx = 4 + i
    data_r = Reference(ws_chart, min_col=col_idx, max_col=col_idx, min_row=3, max_row=12)
    series = Series(data_r, title=lbl)
    series.graphicalProperties.solidFill = color
    bar2.series.append(series)

cats2 = Reference(ws_chart, min_col=2, min_row=4, max_row=13)
bar2.set_categories(cats2)
ws_chart.add_chart(bar2, "J33")

# ── Sheet 4: バックテスト & 論文解説 ─────────────────────────
ws_paper = wb.create_sheet("📚論文・用語解説")
ws_paper.sheet_view.showGridLines = False
ws_paper.column_dimensions["A"].width = 2
ws_paper.column_dimensions["B"].width = 22
ws_paper.column_dimensions["C"].width = 60
ws_paper.column_dimensions["D"].width = 20

ws_paper.merge_cells("B1:D1")
c = ws_paper["B1"]
c.value = "ニューロファイナンス 論文→設計マッピング"
c.font = Font(name="Meiryo UI", bold=True, size=14, color=C_NAVY)
c.alignment = align("left")
ws_paper.row_dimensions[1].height = 32

papers = [
    ("[P1] Kuhnen & Knutson (2005)", "NAcc / Insula 二重回路",
     "報酬期待でNAccが活性 → 利回り得点の上限（8%超でペナルティ）\n損失回避でInsulaが活性 → 財務健全性スコアで発動"),
    ("[P2] Tom et al. (2007)", "λ=2.0 損失回避係数",
     "損失の不快感 = 同額の利益の喜びの2倍\n→ 負のモメンタム・高配当性向は2倍の重みでペナルティ"),
    ("[P3] Knutson & Bossaerts (2007)", "NAcc / Insula 独立性",
     "両回路は独立して作動 → 安定性(Insula)と利回り得点(NAcc)を別軸で設計\n合算しない（独立ウェイト）"),
    ("[P4] Schultz et al. (1997)", "ドーパミン予測誤差",
     "予測より良い結果 → ドーパミン↑（モメンタムスコア）\n連続増配 → 毎期の正の予測誤差 → streakスコア（対数スケール）"),
    ("[P5] Frydman & Camerer (2016)", "処分効果・参照点",
     "減配は参照点を下回る損失体験 → has_dividend_cut で安定性を0.05にキャップ"),
]
ws_paper.cell(2, 2, "論文・コード").font = Font(bold=True, size=9, color=C_WHITE)
ws_paper.cell(2, 2).fill = fill(C_NAVY)
ws_paper.cell(2, 3, "設計への反映").font = Font(bold=True, size=9, color=C_WHITE)
ws_paper.cell(2, 3).fill = fill(C_NAVY)
ws_paper.cell(2, 4, "論文情報").font = Font(bold=True, size=9, color=C_WHITE)
ws_paper.cell(2, 4).fill = fill(C_NAVY)
for c_ in [ws_paper.cell(2,2), ws_paper.cell(2,3), ws_paper.cell(2,4)]:
    c_.alignment = align()
    c_.border = border
ws_paper.row_dimensions[2].height = 22

for ri, (code, short, detail) in enumerate(papers, start=3):
    ws_paper.cell(ri, 2, code).font = Font(name="Meiryo UI", bold=True, size=9, color=C_NAVY)
    ws_paper.cell(ri, 2).fill = fill(C_LIGHT)
    ws_paper.cell(ri, 2).border = border
    ws_paper.cell(ri, 2).alignment = align("left")
    ws_paper.cell(ri, 3, detail).font = Font(name="Meiryo UI", size=9)
    ws_paper.cell(ri, 3).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws_paper.cell(ri, 3).border = border
    ws_paper.cell(ri, 3).fill = fill(C_WHITE)
    ws_paper.cell(ri, 4, short).font = Font(name="Meiryo UI", size=9, color="6B7280")
    ws_paper.cell(ri, 4).alignment = align("left")
    ws_paper.cell(ri, 4).border = border
    ws_paper.row_dimensions[ri].height = 52

# 重みテーブル
ws_paper.row_dimensions[9].height = 16
ws_paper.merge_cells("B10:D10")
c = ws_paper["B10"]
c.value = "▼ スコアウェイト（論文根拠）"
c.font = Font(name="Meiryo UI", bold=True, size=11, color=C_WHITE)
c.fill = fill("374151")
c.alignment = align("left")
ws_paper.row_dimensions[10].height = 24

weights = [
    ("安定性（配当変動CV）", "25%", "[P3] Insula 独立回路"),
    ("財務健全性（自己資本・D/E）", "20%", "[P1] Insula 誘発予防"),
    ("配当利回り（NAcc最適ゾーン）", "20%", "[P1][P3] NAcc 期待報酬"),
    ("配当モメンタム（3年CAGR）", "15%", "[P4] ドーパミン予測誤差"),
    ("配当性向（持続可能性）", "10%", "[P2] λ=2 過剰性向ペナルティ"),
    ("連続増配年数", "7%", "[P4] 連続正予測誤差"),
    ("セクター内順位", "3%", "社会比較参照点"),
]
for ri, (axis, weight, paper_ref) in enumerate(weights, start=11):
    ws_paper.cell(ri, 2, axis).font = Font(name="Meiryo UI", size=9); ws_paper.cell(ri, 2).border = border; ws_paper.cell(ri, 2).alignment = align("left")
    ws_paper.cell(ri, 3, weight).font = Font(name="Meiryo UI", size=9, bold=True, color=C_BLUE); ws_paper.cell(ri, 3).border = border; ws_paper.cell(ri, 3).alignment = align()
    ws_paper.cell(ri, 4, paper_ref).font = Font(name="Meiryo UI", size=9, color="6B7280"); ws_paper.cell(ri, 4).border = border; ws_paper.cell(ri, 4).alignment = align("left")
    ws_paper.row_dimensions[ri].height = 18
    bg = C_LIGHT if ri % 2 == 0 else C_WHITE
    for ci_ in [2, 3, 4]:
        ws_paper.cell(ri, ci_).fill = fill(bg)

# ── 保存 ─────────────────────────────────────────────────────
wb.save(DST)
print(f"\n✅ 完了: {DST}")
print(f"✅ JSON: {JSON_OUT}")
